import os
import sqlite3
import sys 
from flask import (
    Flask, current_app, flash, render_template, request, redirect, url_for, session, g,
    jsonify, send_file 
)
from uuid import uuid4 
from io import BytesIO
import openpyxl 
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

try:
    from models import connect_db, init_db, get_next_barcode, create_discount, get_all_discounts, get_discount, get_next_bill_number
except ImportError:
    print("Error: models.py not found or missing required functions (connect_db, init_db, get_next_barcode, create_discount, get_all_discounts, get_discount, get_next_bill_number).")
    exit(1)

try:
    from utils import generate_barcode, get_sales_analytics, get_custom_range_analytics 
except ImportError:
    print("Error: utils.py not found or missing required functions (generate_barcode, get_sales_analytics, get_custom_range_analytics).")
    exit(1)

from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import traceback # For detailed error logging

# --- PyInstaller Resource Path ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

if getattr(sys, 'frozen', False):
    template_dir = resource_path('templates')
    static_dir = resource_path('static')
    app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
else:
    app = Flask(__name__)

# --- App Configuration ---
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'your_super_secret_key_CHANGE_THIS_LATER_!@#$')
app.permanent_session_lifetime = timedelta(hours=8)
app.config['SESSION_TYPE'] = 'filesystem'

# --- Database Setup ---
def get_database_path_for_init_check():
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        persistent_data_dir = os.path.dirname(sys.executable)
        return os.path.join(persistent_data_dir, 'instance', 'toystore.db')
    else:
        return os.path.join(os.path.abspath("."), 'instance', 'toystore.db')

DATABASE_FOR_CHECK = get_database_path_for_init_check()

if not os.path.exists(DATABASE_FOR_CHECK):
    try:
        print(f"Database not found at {DATABASE_FOR_CHECK}, initializing...")
        os.makedirs(os.path.dirname(DATABASE_FOR_CHECK), exist_ok=True)
        init_db() 
        print("Database initialized.")
    except Exception as e:
        print(f"Error initializing database: {e}")
        traceback.print_exc()
        exit(1)

def get_db():
    if not hasattr(g, 'sqlite_db'):
        g.sqlite_db = connect_db()
        g.sqlite_db.row_factory = sqlite3.Row
    return g.sqlite_db

@app.before_request
def before_request_db(): 
    try:
        get_db()
    except sqlite3.Error as e:
        print(f"Database connection error in before_request: {e}")
        traceback.print_exc()
        flash("Database connection failed. Please try again later.", "danger")

@app.teardown_request
def teardown_request_db(exception): 
    db = getattr(g, 'sqlite_db', None)
    if db is not None:
        db.close()

@app.context_processor
def inject_user_role_now():
    return dict(
        user=session.get('user'),
        role=session.get('role'),
        now=datetime.utcnow()
    )

# --- Core Routes (Auth, Dashboard) ---
@app.route('/')
def home():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    db = get_db()
    if db is None: flash("Database unavailable for registration.", "danger"); return render_template('register.html')
    allow_registration = False
    try:
        cur = db.cursor()
        cur.execute("SELECT COUNT(*) FROM users")
        user_count = cur.fetchone()[0]
        if user_count == 0 or ('user' in session and session.get('role') == 'admin'):
            allow_registration = True
    except sqlite3.Error as e:
        flash(f"Database error checking user count: {e}", "danger")
        traceback.print_exc()

    if not allow_registration:
         flash("Registration is currently restricted.", "warning")
         if request.method == 'POST': return redirect(url_for('login'))
         return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password_input = request.form.get('password', '')
        role = request.form.get('role')
        if not username or not password_input or role not in ['admin', 'cashier']:
            flash("Username, password, and a valid role (Admin/Cashier) are required.", "danger")
            return render_template('register.html')
        
        cur = db.cursor()
        try:
            cur.execute("SELECT id FROM users WHERE username=?", (username,))
            if cur.fetchone():
                flash(f"Username '{username}' already exists. Please choose another.", "warning")
                return render_template('register.html')
            
            password_hash = generate_password_hash(password_input)
            cur.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                      (username, password_hash, role))
            db.commit()
            flash("User registered successfully. Please login.", "success")
            return redirect(url_for('login'))
        except sqlite3.Error as e:
            db.rollback()
            flash(f"Database error during registration: {e}", "danger")
            traceback.print_exc()
        except Exception as e:
             flash(f"An unexpected error occurred during registration: {e}", "danger")
             traceback.print_exc()
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password_input = request.form.get('password', '')
        if not username or not password_input:
             flash("Username and password are required.", "warning")
             return render_template('login.html')
        
        db = get_db()
        if db is None: flash("Database unavailable for login.", "danger"); return render_template('login.html')
        
        cur = db.cursor()
        try:
            cur.execute("SELECT id, username, password, role FROM users WHERE username=?", (username,))
            user = cur.fetchone()
            if user and check_password_hash(user['password'], password_input):
                session.permanent = True
                session['user'] = user['username']
                session['user_id'] = user['id'] 
                session['role'] = user['role']
                flash(f"Welcome back, {user['username']}!", "success")
                return redirect(url_for('dashboard'))
            else:
                flash("Login failed: Incorrect username or password.", "danger")
        except sqlite3.Error as e:
             flash(f"Database error during login: {e}", "danger")
             traceback.print_exc()
        except Exception as e:
             flash(f"An unexpected error occurred during login: {e}", "danger")
             traceback.print_exc()
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for('login'))

@app.route('/dashboard', methods=['GET']) 
def dashboard():
    if 'user' not in session:
        flash("Please log in to access the dashboard.", "warning")
        return redirect(url_for('login'))
    
    db = get_db()
    if db is None:
        flash("Database connection is unavailable. Analytics cannot be displayed.", "danger")
        return render_template('dashboard.html', best_sellers=[], unsold_products=[], daily_sales=[], 
                               weekly_sales=[], monthly_sales=[], yearly_sales=[],
                               custom_range_analytics=None, from_date_filter=None, to_date_filter=None)
    
    best_sellers = get_sales_analytics(db, 'best_sellers')
    unsold_products = get_sales_analytics(db, 'unsold')
    daily_sales = get_sales_analytics(db, 'daily')
    weekly_sales = get_sales_analytics(db, 'weekly')
    monthly_sales = get_sales_analytics(db, 'monthly')
    yearly_sales = get_sales_analytics(db, 'yearly')

    from_date_filter = request.args.get('from_date')
    to_date_filter = request.args.get('to_date')
    custom_range_analytics = None

    if from_date_filter and to_date_filter:
        try:
            from_dt = datetime.strptime(from_date_filter, '%Y-%m-%d')
            to_dt = datetime.strptime(to_date_filter, '%Y-%m-%d')
            if from_dt > to_dt:
                flash("From Date cannot be after To Date for custom analytics.", "warning")
            else:
                custom_range_analytics = get_custom_range_analytics(db, from_date_filter, to_date_filter)
                if custom_range_analytics and custom_range_analytics.get("error"):
                    flash(f"Error fetching custom range analytics: {custom_range_analytics['error']}", "danger")
                    custom_range_analytics = None 
        except ValueError:
            flash("Invalid date format for custom analytics. Please use YYYY-MM-DD.", "warning")
        except Exception as e:
            flash(f"Error processing custom date range: {e}", "danger")
            traceback.print_exc()
            custom_range_analytics = None

    return render_template(
        'dashboard.html',
        best_sellers=best_sellers,
        unsold_products=unsold_products,
        daily_sales=daily_sales,
        weekly_sales=weekly_sales,
        monthly_sales=monthly_sales,
        yearly_sales=yearly_sales,
        custom_range_analytics=custom_range_analytics,
        from_date_filter=from_date_filter, 
        to_date_filter=to_date_filter     
    )

@app.route('/export-dashboard-custom-range')
def export_dashboard_custom_range():
    if 'user' not in session:
        flash("Please log in to export data.", "warning")
        return redirect(url_for('login'))

    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')

    if not from_date_str or not to_date_str:
        flash("Date range is required for export.", "warning")
        return redirect(url_for('dashboard'))

    db = get_db()
    if db is None:
        flash("Database connection unavailable for export.", "danger")
        return redirect(url_for('dashboard'))

    try:
        analytics_data = get_custom_range_analytics(db, from_date_str, to_date_str)
        if not analytics_data or analytics_data.get("error"):
            flash(f"Could not fetch data for export: {analytics_data.get('error', 'Unknown error')}", "danger")
            return redirect(url_for('dashboard', from_date=from_date_str, to_date=to_date_str))

        wb = openpyxl.Workbook()
        
        ws_summary = wb.active
        ws_summary.title = "Custom Range Summary"
        headers_summary = [
            ("Metric", "Value"),
            ("Date Range", f"{from_date_str} to {to_date_str}"),
            ("Total Gross Sales (₹)", f"{analytics_data.get('total_gross_sales', 0):.2f}"),
            ("Total Discounts on Bills (₹)", f"{analytics_data.get('total_discounts_on_bills', 0):.2f}"),
            ("Total Returns Value (₹)", f"{analytics_data.get('total_returns_value', 0):.2f}"),
            ("Total Net Sales (₹)", f"{analytics_data.get('total_net_sales', 0):.2f}"),
            # Profit related metrics for export
            ("Total Cost of Gross Sales (₹)", f"{analytics_data.get('total_cost_of_gross_sales', 0):.2f}"),
            ("Total Cost of Returned Goods (₹)", f"{analytics_data.get('total_cost_of_returned_goods', 0):.2f}"),
            ("Total Net COGS (₹)", f"{analytics_data.get('total_net_cogs', 0):.2f}"),
            ("Estimated Total Profit (₹)", f"{analytics_data.get('estimated_total_profit', 0):.2f}"),
            # End profit related metrics
            ("Number of Bills/Transactions", analytics_data.get('number_of_bills', 0)),
            ("Total Items Sold (Gross Qty)", analytics_data.get('total_items_sold_gross_qty', 0)),
            ("Avg. Items per Bill", f"{analytics_data.get('avg_items_per_bill', 0):.2f}"),
            ("Avg. Bill Value (Net, ₹)", f"{analytics_data.get('avg_bill_value_net', 0):.2f}"),
        ]
        for row_idx, (header, value) in enumerate(headers_summary, 1):
            ws_summary.cell(row=row_idx, column=1, value=header).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        ws_top_qty = wb.create_sheet(title="Top Sellers (Net Qty)")
        ws_top_qty.append(["Product ID", "Product Name", "Gross Sold Qty", "Returned Qty", "Net Sold Qty"])
        for cell in ws_top_qty[1]: cell.font = Font(bold=True)
        for item in analytics_data.get('top_sellers_by_net_qty', []):
            ws_top_qty.append([item['id'], item['name'], item['gross_sold_qty'], item['returned_qty'], item['net_sold_qty']])
        for i, col_letter in enumerate(['A', 'B', 'C', 'D', 'E']): 
            ws_top_qty.column_dimensions[col_letter].width = 15 if col_letter not in ['B'] else 35
        
        ws_top_rev = wb.create_sheet(title="Top Sellers (Gross Rev)")
        ws_top_rev.append(["Product ID", "Product Name", "Gross Revenue (₹)", "Gross Sold Qty", "Returned Qty"])
        for cell in ws_top_rev[1]: cell.font = Font(bold=True)
        for item in analytics_data.get('top_sellers_by_gross_revenue', []):
            ws_top_rev.append([item['id'], item['name'], f"{item['gross_revenue']:.2f}", item['gross_sold_qty'], item['returned_qty']])
        for i, col_letter in enumerate(['A', 'B', 'C', 'D', 'E']): 
            ws_top_rev.column_dimensions[col_letter].width = 20 if col_letter not in ['B'] else 35
            if col_letter == 'C': ws_top_rev.column_dimensions[col_letter].number_format = '"₹"#,##0.00'

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return send_file(
            excel_buffer, as_attachment=True,
            download_name=f"custom_analytics_{from_date_str}_to_{to_date_str}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f"Error exporting custom range data: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('dashboard', from_date=from_date_str, to_date=to_date_str))

# --- Product Management Routes ---
@app.route('/add-product', methods=['GET', 'POST'])
def add_product():
    if 'user' not in session:
        flash("Please log in.", "warning")
        return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database unavailable.", "danger"); return render_template('add_product.html', categories=[])
    
    cur = db.cursor()
    categories = []
    try:
        cur.execute("SELECT id, name FROM categories ORDER BY name")
        categories = cur.fetchall()
    except sqlite3.Error as e: 
        flash(f"Could not load categories: {e}", "danger")
        traceback.print_exc()
        categories = []

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        cost_price_str = request.form.get('cost_price', '0.0')
        selling_price_str = request.form.get('selling_price', '0.0')
        qty_str = request.form.get('quantity', '0')
        cat_id_str = request.form.get('category_id')
        code = request.form.get('barcode', '').strip()
        
        errors = []
        cost_price, selling_price, qty, cat_id = 0.0, 0.0, 0, None

        if not name: errors.append("Product name is required.")
        try: 
            cost_price = float(cost_price_str)
            if cost_price < 0: errors.append("Cost price cannot be negative.")
        except ValueError: errors.append("Invalid cost price format.")
        
        try: 
            selling_price = float(selling_price_str)
            if selling_price < 0: errors.append("Selling price cannot be negative.")
        except ValueError: errors.append("Invalid selling price format.")
        
        try: 
            qty = int(qty_str)
            if qty < 0: errors.append("Quantity cannot be negative.")
        except ValueError: errors.append("Invalid quantity format.")
        
        try:
            if cat_id_str:
                cat_id = int(cat_id_str)
                cur.execute("SELECT id FROM categories WHERE id = ?", (cat_id,))
                if not cur.fetchone(): errors.append("Selected category does not exist.")
            else:
                errors.append("Category is required.")
        except ValueError: errors.append("Invalid category selected.")
        
        final_code = None
        if not code: 
            try:
                final_code = get_next_barcode()
                cur.execute("SELECT id FROM products WHERE barcode = ?", (final_code,))
                if cur.fetchone(): 
                    errors.append(f"Generated barcode '{final_code}' already exists. Please try saving again or enter a unique one."); final_code = None
            except Exception as e: 
                errors.append(f"Could not auto-generate barcode: {e}"); final_code = None
                traceback.print_exc()
        else: 
             cur.execute("SELECT id FROM products WHERE barcode = ?", (code,))
             if cur.fetchone(): 
                 errors.append(f"Barcode '{code}' already exists. Use a unique barcode or leave blank to auto-generate."); final_code = None
             else: 
                 final_code = code
        
        if errors:
            for error in errors: flash(error, 'danger')
        elif final_code:
            barcode_image_path_base = None
            generated_barcode_file_full_path = None
            try:
                barcode_dir_for_saving = os.path.join(app.static_folder, 'barcodes')
                os.makedirs(barcode_dir_for_saving, exist_ok=True)
                barcode_image_path_base = os.path.join(barcode_dir_for_saving, final_code) 
                generated_barcode_file_full_path = generate_barcode(final_code, barcode_image_path_base)
                
                cur.execute("INSERT INTO products (name, cost_price, selling_price, quantity, category_id, barcode) VALUES (?, ?, ?, ?, ?, ?)", 
                              (name, cost_price, selling_price, qty, cat_id, final_code))
                db.commit()
                flash(f"Product '{name}' added with barcode {final_code}", 'success')
                return redirect(url_for('product_list'))
            except sqlite3.Error as e:
                db.rollback()
                flash(f"Database error adding product: {e}", 'danger')
                traceback.print_exc()
                if generated_barcode_file_full_path and os.path.exists(generated_barcode_file_full_path):
                    try: os.remove(generated_barcode_file_full_path)
                    except OSError as file_err: print(f"Error removing barcode file {generated_barcode_file_full_path} on DB error: {file_err}")
            except Exception as e:
                 flash(f"An unexpected error occurred: {e}", 'danger')
                 traceback.print_exc()
                 if generated_barcode_file_full_path and os.path.exists(generated_barcode_file_full_path):
                    try: os.remove(generated_barcode_file_full_path)
                    except OSError as file_err: print(f"Error removing barcode file {generated_barcode_file_full_path} on general error: {file_err}")
    return render_template('add_product.html', categories=categories)

@app.route('/add-category', methods=['GET', 'POST'])
def add_category():
    if 'user' not in session or session.get('role') != 'admin':
        flash("You do not have permission to add categories.", "danger")
        return redirect(url_for('dashboard'))
    
    db = get_db()
    if db is None: flash("Database unavailable.", "danger"); return render_template('add_category.html')
    
    if request.method == 'POST':
        category_name = request.form.get('category_name', '').strip()
        if not category_name: 
            flash("Category name cannot be empty.", "warning")
        else:
            cur = db.cursor()
            try:
                cur.execute("INSERT INTO categories (name) VALUES (?)", (category_name,))
                db.commit()
                flash(f"Category '{category_name}' added successfully.", "success")
                return redirect(url_for('dashboard')) 
            except sqlite3.IntegrityError: 
                db.rollback()
                flash(f"Category '{category_name}' already exists.", "warning")
            except sqlite3.Error as e: 
                db.rollback()
                flash(f"Database error adding category: {e}", "danger")
                traceback.print_exc()
            except Exception as e: 
                flash(f"An unexpected error occurred: {e}", "danger")
                traceback.print_exc()
    return render_template('add_category.html')

@app.route('/products')
def product_list():
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database connection is unavailable.", "danger"); return render_template('product_list.html', products=[], categories=[], search_term='', selected_category='')
    
    cur = db.cursor()
    products_data = []
    categories_list = []
    search_term = request.args.get('search', '').strip()
    category_filter_str = request.args.get('category', '')
    
    try:
        cur.execute("SELECT id, name FROM categories ORDER BY name"); 
        categories_list = cur.fetchall()
        
        query = """SELECT p.id, p.name, p.cost_price, p.selling_price, p.quantity, p.barcode, 
                          c.name as category_name, c.id as category_id 
                   FROM products p 
                   LEFT JOIN categories c ON p.category_id = c.id"""
        params = []
        conditions = []
        
        if search_term: 
            conditions.append("p.name LIKE ? COLLATE NOCASE")
            params.append(f"%{search_term}%")
        
        if category_filter_str:
            try: 
                selected_category_id = int(category_filter_str)
                conditions.append("p.category_id = ?")
                params.append(selected_category_id)
            except ValueError: 
                flash("Invalid category filter value.", "warning")
                category_filter_str = '' 
        
        if conditions: 
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY p.name COLLATE NOCASE"
        
        cur.execute(query, params)
        for row in cur.fetchall():
            cost = row['cost_price'] if row['cost_price'] is not None else 0.0
            sell = row['selling_price'] if row['selling_price'] is not None else 0.0
            profit = sell - cost
            profit_margin = (profit / sell * 100) if sell != 0 else 0
            
            products_data.append({
                'id': row['id'], 'name': row['name'], 'cost_price': cost, 'selling_price': sell, 
                'profit': round(profit, 2), 'profit_margin': round(profit_margin, 2),
                'quantity': row['quantity'] if row['quantity'] is not None else 0, 
                'barcode': row['barcode'], 'category_name': row['category_name'] or 'Uncategorized', 
                'category_id': row['category_id']
            })
    except sqlite3.Error as e: 
        flash(f"Error fetching product list: {e}", "danger")
        traceback.print_exc()
        products_data = []; categories_list = []
        
    return render_template('product_list.html', 
                           products=products_data, categories=categories_list, 
                           search_term=search_term, selected_category=category_filter_str)

@app.route('/edit-product/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database unavailable.", "danger"); return redirect(url_for('product_list'))
    
    cur = db.cursor()
    product = None
    categories_list = []

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        cost_price_str = request.form.get('cost_price', '0.0')
        selling_price_str = request.form.get('selling_price', '0.0')
        qty_str = request.form.get('quantity', '0')
        cat_id_str = request.form.get('category_id')
        
        errors = []
        cost_price, selling_price, quantity, category_id = 0.0, 0.0, 0, None

        if not name: errors.append("Product name is required.")
        try: 
            cost_price = float(cost_price_str)
            if cost_price < 0: errors.append("Cost price cannot be negative.")
        except ValueError: errors.append("Invalid cost price format.")
        
        try: 
            selling_price = float(selling_price_str)
            if selling_price < 0: errors.append("Selling price cannot be negative.")
        except ValueError: errors.append("Invalid selling price format.")
        
        try: 
            quantity = int(qty_str)
            if quantity < 0: errors.append("Quantity cannot be negative.")
        except ValueError: errors.append("Invalid quantity format.")
        
        try:
            if cat_id_str:
                category_id = int(cat_id_str)
                cur.execute("SELECT id FROM categories WHERE id = ?", (category_id,))
                if not cur.fetchone(): errors.append("Selected category does not exist.")
            else:
                errors.append("Category is required.") 
        except ValueError: errors.append("Invalid category selected.")
        
        if errors:
            for error in errors: flash(error, 'danger')
            try:
                cur.execute("SELECT * FROM products WHERE id=?", (product_id,))
                product = cur.fetchone()
                cur.execute("SELECT id, name FROM categories ORDER BY name")
                categories_list = cur.fetchall()
                if not product: 
                    flash("Product not found (while reloading form).", "warning")
                    return redirect(url_for('product_list'))
                return render_template('edit_product.html', product=product, categories=categories_list)
            except Exception as fetch_err: 
                flash(f"Error reloading form data after validation error: {fetch_err}", "danger")
                traceback.print_exc()
                return redirect(url_for('product_list'))
        else: 
            try:
                cur.execute("UPDATE products SET name=?, cost_price=?, selling_price=?, quantity=?, category_id=? WHERE id=?", 
                              (name, cost_price, selling_price, quantity, category_id, product_id))
                db.commit()
                flash("Product updated successfully", "success")
                return redirect(url_for('product_list'))
            except sqlite3.Error as e: 
                db.rollback()
                flash(f"Database error updating product: {e}", "danger")
                traceback.print_exc()
                try:
                    cur.execute("SELECT * FROM products WHERE id=?", (product_id,))
                    product = cur.fetchone() 
                    cur.execute("SELECT id, name FROM categories ORDER BY name")
                    categories_list = cur.fetchall()
                    if not product:
                         flash("Product not found (after DB error).", "warning")
                         return redirect(url_for('product_list'))
                    return render_template('edit_product.html', product=product, categories=categories_list)
                except Exception as reload_err:
                    flash(f"Additional error reloading form after DB update error: {reload_err}", "danger")
                    return redirect(url_for('product_list'))

            except Exception as e: 
                flash(f"An unexpected error occurred during product update: {e}", "danger")
                traceback.print_exc()
    
    try:
        cur.execute("SELECT * FROM products WHERE id=?", (product_id,))
        product = cur.fetchone()
        if not product: 
            flash("Product not found.", "warning")
            return redirect(url_for('product_list'))
        
        cur.execute("SELECT id, name FROM categories ORDER BY name")
        categories_list = cur.fetchall()
    except sqlite3.Error as e: 
        flash(f"Error fetching product details for editing: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('product_list')) 
    except Exception as e: 
        flash(f"An unexpected error occurred fetching product for edit: {e}", "danger")
        traceback.print_exc()
        return redirect(url_for('product_list')) 
    
    return render_template('edit_product.html', product=product, categories=categories_list)


@app.route('/delete-product/<int:product_id>', methods=['POST'])
def delete_product(product_id):
    if 'user' not in session: flash("Authentication required.", "warning"); return redirect(url_for('login'))
    if session.get('role') != 'admin': flash("Permission denied.", "danger"); return redirect(url_for('product_list'))
    
    db = get_db(); 
    if db is None: flash("Database unavailable.", "danger"); return redirect(url_for('product_list'))
    cur = db.cursor(); product_barcode = None
    try:
        cur.execute("SELECT id, name, barcode FROM products WHERE id=?", (product_id,)); product_row = cur.fetchone()
        if not product_row: flash("Product not found.", "warning"); return redirect(url_for('product_list'))
        
        product_name, product_barcode = product_row['name'], product_row['barcode']
        cur.execute("SELECT COUNT(*) FROM sales WHERE product_id = ?", (product_id,)); sales_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM returns WHERE product_id = ?", (product_id,)); returns_count = cur.fetchone()[0]
        
        if sales_count > 0 or returns_count > 0:
            flash(f"Cannot delete '{product_name}': It has associated sales ({sales_count}) or return ({returns_count}) records.", "warning")
            return redirect(url_for('product_list'))
        
        cur.execute("DELETE FROM products WHERE id=?", (product_id,)); db.commit()
        if product_barcode:
            try:
                barcode_path = os.path.join(app.static_folder, 'barcodes', f"{product_barcode}.png")
                if os.path.exists(barcode_path): os.remove(barcode_path); print(f"Deleted barcode image: {barcode_path}")
            except Exception as e: print(f"Warning: Could not delete barcode file {product_barcode}.png: {e}")
        flash(f"Product '{product_name}' deleted successfully.", "success")
    except sqlite3.Error as e: db.rollback(); flash(f"Database error deleting product: {e}", "danger"); traceback.print_exc()
    except Exception as e: flash(f"An unexpected error occurred: {e}", "danger"); traceback.print_exc()
    return redirect(url_for('product_list'))

# --- Billing and Cart Routes ---
import json 
if os.name == 'nt': 
    try: import win32print
    except ImportError: print("Warning: win32print module not found. Thermal printing will be disabled."); win32print = None 
else: win32print = None 


@app.route('/billing', methods=['GET', 'POST'])
def billing():
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database connection unavailable for billing.", "danger"); return render_template('billing.html', cart=[], total=0, products=[], discounts=[], applied_discount_details=None, current_discount_amount_for_display=0.0, total_before_discount=0.0, win32print_available=(win32print is not None))
    if 'cart' not in session: session['cart'] = []
    
    discounts_list = get_all_discounts() 
    applied_discount_id = session.get('applied_bill_discount_id') 
    applied_discount_details = None
    current_discount_amount_for_display = 0.0
    
    cart_items_for_display = session.get('cart', []) 
    total_before_discount_for_display = sum(item['price'] * item['qty'] for item in cart_items_for_display) 

    if applied_discount_id:
        disc_info_tuple = get_discount(int(applied_discount_id)) 
        if disc_info_tuple:
            applied_discount_details = {'id': disc_info_tuple[0], 'name': disc_info_tuple[1], 'percent': disc_info_tuple[2]}
            current_discount_amount_for_display = total_before_discount_for_display * (applied_discount_details['percent'] / 100)

    final_total_payable_for_display = total_before_discount_for_display - current_discount_amount_for_display

    action = None
    if request.method == 'POST':
        if 'barcode' in request.form: action = 'add_scan'
        elif 'select_product_id' in request.form: action = 'add_manual'
        elif 'checkout_and_print' in request.form: action = 'checkout_and_print' 
        elif 'apply_discount' in request.form: action = 'apply_discount'
        elif 'clear_discount' in request.form: action = 'clear_discount'

        try:
            if action == 'add_scan' or action == 'add_manual':
                product_to_add = None
                if action == 'add_scan':
                    barcode = request.form.get('barcode', '').strip()
                    if barcode:
                        cur = db.cursor(); cur.execute("SELECT id, name, selling_price, barcode FROM products WHERE barcode=? AND quantity > 0", (barcode,)); product_to_add = cur.fetchone()
                        if not product_to_add: flash(f"Barcode '{barcode}' not found or out of stock.", "warning")
                    else: flash("Barcode cannot be empty.", "warning")
                else: 
                    pid_str = request.form.get('select_product_id')
                    try:
                        pid = int(pid_str); cur = db.cursor(); cur.execute("SELECT id, name, selling_price, barcode FROM products WHERE id=? AND quantity > 0", (pid,)); product_to_add = cur.fetchone()
                        if not product_to_add: flash(f"Product ID {pid} not found or out of stock.", "warning")
                    except (ValueError, TypeError): flash("Invalid product selection.", "danger")

                if product_to_add:
                    cart_session = session.get('cart', []) 
                    item_found = next((item for item in cart_session if item['id'] == product_to_add['id']), None)
                    if item_found: item_found['qty'] += 1
                    else: cart_session.append({'id': product_to_add['id'], 'barcode': product_to_add['barcode'], 'name': product_to_add['name'], 'price': float(product_to_add['selling_price']), 'qty': 1})
                    session['cart'] = cart_session; session.modified = True; flash(f"Added '{product_to_add['name']}'.", "success")
                return redirect(url_for('billing'))

            elif action == 'apply_discount':
                did_str = request.form.get('discount_id')
                if did_str:
                    session['applied_bill_discount_id'] = did_str 
                    flash(f"Discount selected. Will be applied at checkout.", "info")
                else: 
                    if 'applied_bill_discount_id' in session: del session['applied_bill_discount_id']
                    flash("Discount cleared.", "info")
                session.modified = True
                return redirect(url_for('billing')) 
            
            elif action == 'clear_discount':
                if 'applied_bill_discount_id' in session: del session['applied_bill_discount_id']
                session.modified = True
                flash("Discount cleared.", "info")
                return redirect(url_for('billing'))

            elif action == 'checkout_and_print':
                cart_for_sale = session.get('cart', []) 
                if not cart_for_sale:
                    flash("Cart is empty. Cannot checkout.", "warning")
                    return redirect(url_for('billing'))
                
                payment_method_from_form = request.form.get('payment_method')
                if not payment_method_from_form:
                    flash("Please select a payment method before checking out.", "warning")
                    return redirect(url_for('billing'))
                
                bill_total_gross = sum(item['price'] * item['qty'] for item in cart_for_sale)
                bill_discount_id_final = session.get('applied_bill_discount_id')
                bill_discount_amount_final = 0.0
                applied_discount_percent_for_receipt = 0.0
                
                if bill_discount_id_final:
                    disc_details_tuple = get_discount(int(bill_discount_id_final)) 
                    if disc_details_tuple:
                        applied_discount_percent_for_receipt = disc_details_tuple[2]
                        bill_discount_amount_final = bill_total_gross * (applied_discount_percent_for_receipt / 100)
                
                final_amount_payable = bill_total_gross - bill_discount_amount_final
                sale_time = datetime.now()
                next_bill_no_int = get_next_bill_number()
                bill_identifier_for_db = str(next_bill_no_int) 
                user_id = session.get('user_id')

                receipt_items_data_for_print = []
                for item_in_cart in cart_for_sale:
                    receipt_items_data_for_print.append(
                        (item_in_cart['name'], item_in_cart['qty'], item_in_cart['price'], item_in_cart['price'] * item_in_cart['qty'])
                    )
                
                try:
                    cur = db.cursor()
                    db.execute('BEGIN') 
                    for i, item_processed in enumerate(cart_for_sale):
                        cur.execute("SELECT quantity FROM products WHERE id = ?", (item_processed['id'],))
                        stock = cur.fetchone()
                        if not stock or stock['quantity'] < item_processed['qty']:
                            db.rollback(); flash(f"Insufficient stock for '{item_processed['name']}'. Sale aborted.", "danger"); return redirect(url_for('billing'))
                        
                        current_item_bill_discount = bill_discount_amount_final if i == 0 else 0.0 
                        cur.execute(
                            """INSERT INTO sales (bill_identifier, product_id, quantity, sale_price, sale_date, user_id, discount_applied_to_bill, payment_method) 
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                            (bill_identifier_for_db, item_processed['id'], item_processed['qty'], item_processed['price'], 
                             sale_time, user_id, current_item_bill_discount, payment_method_from_form)
                        )
                        cur.execute("UPDATE products SET quantity = quantity - ? WHERE id = ?", (item_processed['qty'], item_processed['id']))
                    db.commit()
                    
                    sale_flash_msg = f"Sale completed! Bill No: {bill_identifier_for_db}. Total: ₹{final_amount_payable:.2f}. Payment: {payment_method_from_form}."
                    if bill_discount_amount_final > 0: sale_flash_msg += f" (Discount of {applied_discount_percent_for_receipt}% [₹{bill_discount_amount_final:.2f}] applied)."
                    flash(sale_flash_msg, "success")

                    print_success_flag = False; print_status_message = "Printing skipped or printer not available."
                    if win32print: 
                        store_name = "Z Toys And Gifts"; insta_id = "ztoysandgifts"; contact = "7708159325"; address = "MS Road, Parvathipuram, Nagercoil-629003"
                        print_success_flag, print_status_message = print_thermal_receipt(
                            store_name, insta_id, contact, address, receipt_items_data_for_print, 
                            final_amount_payable, bill_discount_amount_final, bill_id=bill_identifier_for_db,
                            payment_method=payment_method_from_form
                        )
                    
                    if print_success_flag: flash(f"Receipt for Bill No: {bill_identifier_for_db} printed successfully.", "info")
                    else: flash(f"Receipt printing for Bill No: {bill_identifier_for_db} status: {print_status_message}", "warning")
                    
                    session['cart'] = []; 
                    if 'applied_bill_discount_id' in session: del session['applied_bill_discount_id']
                    session.modified = True
                    return redirect(url_for('billing', last_bill_id=bill_identifier_for_db)) 

                except Exception as e:
                    db.rollback(); flash(f"Error during checkout & print: {e}", "danger"); traceback.print_exc(); return redirect(url_for('billing'))
        except Exception as e:
            flash(f"Processing error in billing: {e}", "danger"); traceback.print_exc(); return redirect(url_for('billing'))
    
    cur = db.cursor(); cur.execute("SELECT id, name, selling_price FROM products WHERE quantity > 0 ORDER BY name COLLATE NOCASE"); available_products = cur.fetchall()
    last_bill_id_for_print_option = request.args.get('last_bill_id')

    return render_template('billing.html', 
                           cart=cart_items_for_display, total=final_total_payable_for_display, 
                           products=available_products, discounts=discounts_list, 
                           applied_discount_details=applied_discount_details, 
                           current_discount_amount_for_display=current_discount_amount_for_display, 
                           total_before_discount=total_before_discount_for_display,
                           last_bill_id_for_print=last_bill_id_for_print_option,
                           win32print_available=(win32print is not None))

# --- Cart Item Adjustment Routes ---
@app.route('/increase-cart-item/<int:product_id>', methods=['POST'])
def increase_cart_item(product_id):
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    cart = session.get('cart', [])
    found = False; db = get_db()
    if db is None: flash("Database unavailable.", "danger"); return redirect(url_for('billing'))
    cur = db.cursor()
    for item in cart:
        if item.get('id') == product_id:
            item_name = item.get('name', 'Item')
            try:
                cur.execute("SELECT quantity FROM products WHERE id = ?", (product_id,)); stock_row = cur.fetchone()
                if stock_row and stock_row['quantity'] > item['qty']: item['qty'] += 1; found = True; flash(f"Increased quantity for '{item_name}'.", "info")
                elif stock_row: flash(f"No more '{item_name}', stock ({stock_row['quantity']} available).", "warning")
                else: flash(f"Product ID {product_id} not found in database.", "warning")
            except Exception as e: flash(f"Error checking stock: {e}", "danger"); traceback.print_exc()
            break
    if not found and any(item.get('id') == product_id for item in cart): pass 
    elif not found: flash("Item not found in cart.", "warning")
    session['cart'] = cart; session.modified = True
    return redirect(url_for('billing'))

@app.route('/decrease-cart-item/<int:product_id>', methods=['POST'])
def decrease_cart_item(product_id):
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    cart = session.get('cart', []); item_to_remove_idx = -1; item_name = "Item"; found = False
    for i, item in enumerate(cart):
        if item.get('id') == product_id:
            item_name = item.get('name', 'Item'); item['qty'] -= 1
            if item['qty'] <= 0: item_to_remove_idx = i
            found = True; break
    if item_to_remove_idx != -1: del cart[item_to_remove_idx]; flash(f"Removed '{item_name}' from cart.", "info")
    elif found: flash(f"Decreased quantity for '{item_name}'.", "info")
    if not found: flash("Item not found in cart.", "warning")
    session['cart'] = cart; session.modified = True
    return redirect(url_for('billing'))

@app.route('/remove-cart-item/<int:product_id>', methods=['POST'])
def remove_cart_item(product_id):
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    cart = session.get('cart', []); original_length = len(cart); item_name = "Item"
    for item_lookup in cart:
        if item_lookup.get('id') == product_id: item_name = item_lookup.get('name', 'Item'); break
    cart = [item for item in cart if item.get('id') != product_id]
    if len(cart) < original_length: session['cart'] = cart; session.modified = True; flash(f"Removed '{item_name}' from cart.", "success")
    else: flash("Item not found in cart to remove.", "warning")
    return redirect(url_for('billing'))

# --- Return Order Route ---
@app.route('/return-order', methods=['GET', 'POST'])
def return_order():
    if 'user' not in session: flash("Please log in to process returns.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database connection unavailable.", "danger"); return render_template('return_order.html', products=[])
    
    cur = db.cursor()
    products_list = []
    try:
        cur.execute("SELECT id, name, selling_price FROM products ORDER BY name COLLATE NOCASE")
        products_list = cur.fetchall()
    except sqlite3.Error as e: 
        flash(f"Could not load products for return: {e}", "danger")
        traceback.print_exc()
        products_list = []

    if request.method == 'POST':
        product_id_str = request.form.get('product_id')
        quantity_str = request.form.get('quantity')
        return_price_str = request.form.get('return_price')
        reason = request.form.get('reason', '').strip()
        original_bill_identifier = request.form.get('original_bill_identifier', '').strip() 
        
        errors = []
        product_id, quantity, return_price = None, None, None
        product_data_for_flash = None 

        try:
            product_id = int(product_id_str)
            cur.execute("SELECT name FROM products WHERE id = ?", (product_id,))
            product_data_for_flash = cur.fetchone()
            if not product_data_for_flash: 
                errors.append("Invalid product selected (product not found in DB).")
        except (ValueError, TypeError): 
            errors.append("Product ID must be a valid number.")
        
        try:
            quantity = int(quantity_str)
            if quantity <= 0: errors.append("Quantity must be a positive number.")
        except (ValueError, TypeError): errors.append("Quantity must be a valid integer.")
        
        try:
            return_price = float(return_price_str)
            if return_price < 0: errors.append("Return price cannot be negative.")
        except (ValueError, TypeError): errors.append("Return price must be a valid number.")
        
        if errors:
            for error in errors: flash(error, 'danger')
        else:
            try:
                db.execute('BEGIN') 
                cur.execute(
                    """INSERT INTO returns (product_id, quantity, return_price, reason, return_date, original_bill_identifier) 
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (product_id, quantity, return_price, reason or None, datetime.now(), original_bill_identifier or None)
                )
                cur.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?", (quantity, product_id))
                db.commit()
                flash(f"Return of {quantity} x '{product_data_for_flash['name'] if product_data_for_flash else 'Product'}' processed. Stock updated.", 'success')
                return redirect(url_for('return_order')) 
            except sqlite3.Error as e: 
                db.rollback()
                flash(f"Database error processing return: {e}", 'danger')
                traceback.print_exc()
            except Exception as e: 
                db.rollback() 
                flash(f"An unexpected error occurred during return processing: {e}", 'danger')
                traceback.print_exc()
                
    return render_template('return_order.html', products=products_list)

# --- Product Sales Analytics Routes ---
@app.route('/product-analytics')
def product_analytics_page():
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database connection unavailable.", "danger"); return render_template('product_analytics.html', products_summary=[])
    products_summary = []
    try: products_summary = get_sales_analytics(db, 'product_summary')
    except Exception as e: flash(f"Error fetching product analytics data: {e}", "danger"); traceback.print_exc(); products_summary = []
    return render_template('product_analytics.html', products_summary=products_summary)

@app.route('/export-product-analytics')
def export_product_analytics():
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database connection unavailable for export.", "danger"); return redirect(url_for('product_analytics_page'))
    try:
        products_summary = get_sales_analytics(db, 'product_summary')
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Product Sales Analytics"
        headers = ["ID", "Product Name", "Barcode", "Category", "Cost Price", "Selling Price", "Current Stock", "Gross Units Sold", "Units Returned", "Net Units Sold", "Total Revenue (Gross)", "Est. Profit (Gross)"]
        ws.append(headers)
        for product in products_summary:
            ws.append([ product['id'], product['name'], product['barcode'], product['category_name'], product['cost_price'], product['selling_price'], product['current_stock'],
                        product['total_units_sold_gross'], product['total_units_returned'], product['net_units_sold'], product['total_revenue_generated'], product['profit_generated_gross'] ])
        for cell in ws[1]: cell.font = Font(bold=True)
        for col_idx, column_cells in enumerate(ws.columns):
            max_length = 0; column_letter = get_column_letter(col_idx + 1)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50 
            ws.column_dimensions[column_letter].width = adjusted_width
        excel_buffer = BytesIO(); wb.save(excel_buffer); excel_buffer.seek(0)
        return send_file(excel_buffer, as_attachment=True, download_name='product_sales_analytics.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e: flash(f"Error exporting data: {e}", "danger"); traceback.print_exc(); return redirect(url_for('product_analytics_page'))

# --- Bill History Route ---
@app.route('/bill-history')
def bill_history():
    if 'user' not in session: flash("Please log in.", "warning"); return redirect(url_for('login'))
    db = get_db()
    if db is None: flash("Database connection unavailable.", "danger"); return render_template('bill_history.html', bills_list=[])
    bills_data = {}
    try:
        cur = db.cursor()
        cur.execute("""
            SELECT s.bill_identifier, s.sale_date, s.quantity, s.sale_price, s.discount_applied_to_bill, s.payment_method,
                   p.name as product_name, u.username as cashier_name
            FROM sales s JOIN products p ON s.product_id = p.id LEFT JOIN users u ON s.user_id = u.id
            ORDER BY s.sale_date DESC, s.bill_identifier DESC, s.id ASC """)
        all_sale_items = cur.fetchall()
        for item_row in all_sale_items:
            bill_id = item_row['bill_identifier']
            if bill_id not in bills_data:
                sale_date_obj = None
                if item_row['sale_date']:
                    try: sale_date_obj = datetime.strptime(str(item_row['sale_date']), '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        try: sale_date_obj = datetime.strptime(str(item_row['sale_date']), '%Y-%m-%d %H:%M:%S')
                        except ValueError: pass
                formatted_date = sale_date_obj.strftime('%d %b %Y, %I:%M %p') if sale_date_obj else 'N/A'
                bills_data[bill_id] = {
                    'bill_identifier': bill_id, 'sale_date_raw': sale_date_obj, 'sale_date_formatted': formatted_date,
                    'cashier_name': item_row['cashier_name'] or 'N/A', 'bill_products': [],
                    'bill_total_gross': 0.0, 'discount_on_bill': float(item_row['discount_applied_to_bill'] or 0.0),
                    'payment_method': item_row['payment_method'] or 'N/A', 'bill_final_amount': 0.0, 'returns': []
                }
            item_subtotal = item_row['quantity'] * item_row['sale_price']
            bills_data[bill_id]['bill_products'].append({'product_name': item_row['product_name'], 'quantity': item_row['quantity'], 'sale_price': item_row['sale_price'], 'subtotal': item_subtotal})
            bills_data[bill_id]['bill_total_gross'] += item_subtotal
        for bill_id_key in list(bills_data.keys()): 
            bill = bills_data[bill_id_key]
            bill['bill_final_amount'] = bill['bill_total_gross'] - bill['discount_on_bill']
            cur.execute("SELECT r.product_id, p.name as product_name, r.quantity as returned_quantity, r.return_price, r.return_date, r.reason FROM returns r JOIN products p ON r.product_id = p.id WHERE r.original_bill_identifier = ? ORDER BY r.return_date DESC", (bill_id_key,))
            bill_returns = cur.fetchall()
            if bill_returns:
                bill['returns'] = []
                for ret_row in bill_returns:
                    ret_date_obj = None
                    if ret_row['return_date']:
                        try: ret_date_obj = datetime.strptime(str(ret_row['return_date']), '%Y-%m-%d %H:%M:%S.%f')
                        except ValueError:
                            try: ret_date_obj = datetime.strptime(str(ret_row['return_date']), '%Y-%m-%d %H:%M:%S')
                            except ValueError: pass
                    return_item_dict = dict(ret_row); return_item_dict['return_date_formatted'] = ret_date_obj.strftime('%d %b %Y, %I:%M %p') if ret_date_obj else 'N/A'
                    bill['returns'].append(return_item_dict)
    except sqlite3.Error as e: flash(f"Error fetching bill history: {e}", "danger"); traceback.print_exc(); bills_data = {}
    except Exception as e: flash(f"An unexpected error occurred in bill history: {e}", "danger"); traceback.print_exc(); bills_data = {}
    def sort_key_bill_history(b):
        try: return int(b['bill_identifier']) 
        except ValueError: return float('-inf') 
    sorted_bills_list = sorted(bills_data.values(), key=sort_key_bill_history, reverse=True)
    return render_template('bill_history.html', bills_list=sorted_bills_list)

# --- Error Handlers ---
@app.errorhandler(404)
def page_not_found(e): return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    print(f"Internal Server Error 500: {e}"); traceback.print_exc()
    db = getattr(g, 'sqlite_db', None)
    if db is not None:
        try: db.rollback(); print("DB Rolled back due to 500 error.")
        except Exception as rollback_e: print(f"Error during DB rollback on 500 error: {rollback_e}"); traceback.print_exc()
    flash("An internal server error occurred. Please try again later or contact support.", "danger")
    if 'user' in session: return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# --- Receipt Printing ---
@app.route('/print-receipt', methods=['POST'])
def print_receipt_ajax():
    if 'user' not in session:
        return jsonify({"success": False, "message": "Authentication required."}), 401
    
    if not win32print:
        return jsonify({"success": False, "message": "Printing disabled on this system."}), 500

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"success": False, "message": "Invalid request data."}), 400

    bill_id = data.get('bill_id')
    items_data = data.get('items', [])
    
    if not items_data or not isinstance(items_data, list):
        return jsonify({"success": False, "message": "Invalid or empty items list."}), 400

    # Process items
    items = []
    for i, item in enumerate(items_data):
        try:
            name = str(item.get('name', f'Item #{i+1}')).strip()
            qty = max(1, int(item.get('qty', 1)))
            price = max(0, float(item.get('price', 0)))
            subtotal = item.get('subtotal')
            if subtotal is None:
                subtotal = price * qty
            else:
                subtotal = max(0, float(subtotal))
            items.append((name, qty, price, subtotal))
        except (ValueError, TypeError) as e:
            print(f"Error processing item #{i+1}: {e}")
            continue

    if not items:
        return jsonify({"success": False, "message": "No valid items to print."}), 400

    # Get other receipt data
    total = float(data.get('total', 0))
    discount = float(data.get('discount', 0))
    payment_method = data.get('payment_method', 'Cash')

    # Store details
    store_name = "Z Toys And Gifts"
    insta_id = "ztoysandgifts"
    contact = "7708159325"
    address = "MS Road, Parvathipuram, Nagercoil-629003"

    # Print the receipt
    success, message = print_thermal_receipt(
        store_name, insta_id, contact, address,
        items, total,
        discount_amount=discount,
        bill_id=bill_id,
        payment_method=payment_method
    )

    return jsonify({"success": success, "message": message})


def print_thermal_receipt(store_name, insta_id, contact, address, items, total,
                          discount_amount=0, bill_id=None, payment_method=None,
                          printer_name=None):
    if not win32print:
        return False, "Printing not supported or win32print missing."

    # Choose default printer if none provided
    if printer_name is None:
        try:
            printer_name = win32print.GetDefaultPrinter()
        except Exception as e:
            traceback.print_exc()
            return False, f"Error getting default printer: {e}"

    try:
        h_printer = win32print.OpenPrinter(printer_name)
    except Exception as e:
        traceback.print_exc()
        return False, f"Could not open printer: {e}"

    # ESC/POS commands
    ESC = b'\x1b'
    GS = b'\x1d'
    INIT = ESC + b'@'
    BOLD_ON = ESC + b'E' + b'\x01'
    BOLD_OFF = ESC + b'E' + b'\x00'
    SIZE_BIG = GS + b'!' + b'\x11'
    SIZE_NORMAL = GS + b'!' + b'\x00'
    ALIGN_CENTER = ESC + b'a' + b'\x01'
    ALIGN_LEFT = ESC + b'a' + b'\x00'
    ALIGN_RIGHT = ESC + b'a' + b'\x02'
    FEED_4 = ESC + b'd' + b'\x04'
    CUT = GS + b'V' + b'\x00'
    
    # Start building the receipt
    rc = bytearray()
    rc += INIT

    # Header: store name (centered)
    rc += ALIGN_CENTER + SIZE_BIG + BOLD_ON
    rc += store_name.encode('utf-8', 'replace') + b'\n'
    rc += BOLD_OFF + SIZE_NORMAL

    # Date/time and bill ID (centered)
    now_str = datetime.now().strftime('%d/%m/%Y %I:%M %p')
    rc += ALIGN_CENTER + now_str.encode('utf-8', 'replace') + b'\n'
    
    if bill_id:
        rc += ALIGN_CENTER + f"Bill No: {bill_id}".encode('utf-8', 'replace') + b'\n'

    # Contact info & address (left aligned)
    rc += ALIGN_LEFT
    rc += f"@{insta_id}\n".encode('utf-8', 'replace')
    rc += f"Contact: {contact}\n".encode('utf-8', 'replace')
    rc += address.encode('utf-8', 'replace') + b'\n\n'

    # Items section
    # Separator line
    separator = b'-' * 48 + b'\n'
    rc += separator

    # Column headers
    header_line = f"{'Item':<22} {'Qty':>3} {'Price':>8} {'Total':>10}\n"
    rc += header_line.encode('utf-8', 'replace')
    rc += separator

    # Item lines
    for name, qty, price_per_item, subtotal in items:
        # Truncate name if too long
        display_name = (name[:20] + '..') if len(name) > 22 else name
        
        # Format the line with proper spacing
        item_line = f"{display_name:<22} {qty:>3} {price_per_item:>8.2f} {subtotal:>10.2f}\n"
        rc += item_line.encode('utf-8', 'replace')

    rc += separator

    # Totals section (right aligned)
    if discount_amount > 0:
        original_total = total + discount_amount
        
        # Original total line
        total_line = f"{'Subtotal:':>30} {original_total:>12.2f}\n"
        rc += total_line.encode('utf-8', 'replace')
        
        # Discount line
        discount_line = f"{'Discount:':>30} {-discount_amount:>12.2f}\n"
        rc += discount_line.encode('utf-8', 'replace')
        
        rc += separator

    # Final total (bold)
    rc += BOLD_ON
    final_total_line = f"{'TOTAL:':>30} {total:>12.2f}\n"
    rc += final_total_line.encode('utf-8', 'replace')
    rc += BOLD_OFF

    # Payment method
    if payment_method:
        payment_line = f"{'Payment:':>30} {payment_method:>12}\n"
        rc += payment_line.encode('utf-8', 'replace')

    # Footer
    rc += b'\n'
    rc += ALIGN_CENTER + b"Thank you for shopping!\n"
    rc += b"Have a great day!\n"
    
    # Feed paper and cut
    rc += FEED_4 + CUT

    # Send to printer
    try:
        win32print.StartDocPrinter(h_printer, 1, ("Receipt", None, "RAW"))
        win32print.StartPagePrinter(h_printer)
        win32print.WritePrinter(h_printer, bytes(rc))
        win32print.EndPagePrinter(h_printer)
        win32print.EndDocPrinter(h_printer)
        win32print.ClosePrinter(h_printer)
        return True, "Printed successfully"
    except Exception as e:
        traceback.print_exc()
        try:
            win32print.ClosePrinter(h_printer)
        except:
            pass
        return False, str(e)
@app.route('/available-printers')
def get_printers():
    if 'user' not in session: return jsonify({"success": False, "message": "Log in first"}), 401
    if not win32print: return jsonify({"success": False, "message": "Printing disabled."}), 500
    try:
        printers_list = [p[2] for p in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL, None, 1)]
        default_printer = win32print.GetDefaultPrinter()
        return jsonify({"success": True, "printers": printers_list, "default": default_printer})
    except Exception as e: traceback.print_exc(); return jsonify({"success": False, "message": f"Error fetching printers: {str(e)}"}), 500

#--Discounts--
@app.route('/discounts')
def discount_list():
    if 'user' not in session: return redirect(url_for('login'))
    discounts_data = get_all_discounts() 
    return render_template('discounts.html', discounts=discounts_data)

@app.route('/discounts/new', methods=['GET','POST'])
def discount_new():
    if 'user' not in session or session.get('role')!='admin': flash("Admin only.","danger"); return redirect(url_for('discount_list'))
    if request.method=='POST':
        name = request.form['name'].strip().upper(); percent_str = request.form['percent']
        try:
            percent = float(percent_str)
            if not name or percent < 0 or percent > 100: flash("Invalid discount name or percentage (0-100).","danger")
            else: create_discount(name, percent); flash("Discount created.","success"); return redirect(url_for('discount_list'))
        except ValueError: flash("Invalid percentage format.", "danger")
        except sqlite3.IntegrityError: flash(f"Discount '{name}' already exists or other DB constraint violated.", "danger")
        except Exception as e: flash(f"Could not create discount: {e}","danger"); traceback.print_exc()
    return render_template('discount_new.html')

# --- Main Execution ---
if __name__ == '__main__':
    barcode_dir = os.path.join(app.static_folder, 'barcodes')
    if not os.path.exists(barcode_dir):
        try: os.makedirs(barcode_dir); print(f"Created directory: {barcode_dir}")
        except OSError as e: print(f"Error creating barcode directory {barcode_dir}: {e}"); traceback.print_exc()
    if not os.path.exists(DATABASE_FOR_CHECK):
        print(f"Database still not found at {DATABASE_FOR_CHECK} before run, attempting init again...")
        try:
            os.makedirs(os.path.dirname(DATABASE_FOR_CHECK), exist_ok=True)
            init_db(); print("Database initialized successfully before run.")
        except Exception as e: print(f"CRITICAL: Failed to initialize database before run: {e}"); traceback.print_exc(); exit(1)
    app.run(debug=True, host='0.0.0.0', port=5000)